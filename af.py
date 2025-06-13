import csv
import os
from openpyxl import Workbook
import pandas as pd
import subprocess

average_af_ref_per_row = []
average_af_ref_per_row1 = []
average_af_ref_per_row2 = []
average_af_alt_per_row = []
average_af_alt_per_row1 = []
average_af_alt_per_row2 = []
average_gt_het_per_row = []
average_gt_het_per_row1 = []
average_gt_het_per_row2 = []

#spaincsv = bran_spv.csv E10_exp2v.csv E17_e15v.csv E1_E6_exp2v.csv E20v.csv E21v.csv E22v.csv E2v.csv E3_exp2v.csv E4_exp2_spv.csv E5v.csv est_E18v.csv leit_spainv.csv ord_e19_spv.csv pajaresv.csv pena_e9v.csv PORTv.csv spain_pyrenesv.csv vega_spav.csv yeclav.csv 
#scadcsv = nor1v.csv nor2v.csv nor4v.csv nor5v.csv nor6v.csv nor7v.csv par_nor6v.csv s1v.csv s2v.csv S3v.csv S5v.csv sve2v.csv sve3v.csv sve4v.csv $scad
#alpscsv = austriav.csv bri_frv.csv chv.csv cr_frv.csv D1_D2v.csv D3v.csv D4v.csv Dotherv.csv Fgal1v.csv fr1_5v.csv fr4v.csv fr6v.csv fr_Pv.csv galiber_frv.csv ganon_frv.csv gv_frv.csv polandv.csv vallon_frv.csv $alps

script_directory = os.path.dirname(os.path.abspath(__file__))

spaincsv = ['bran_spv.csv', 'E10_exp2v.csv', 'E17_e15v.csv', 'E1_E6_exp2v.csv', 'E20v.csv', 'E21v.csv', 'E22v.csv', 'E2v.csv', 'E3_exp2v.csv', 'E4_exp2_spv.csv',
'E5v.csv', 'est_E18v.csv', 'leit_spainv.csv','ord_e19_spv.csv','pajaresv.csv','pena_e9v.csv','PORTv.csv','spain_pyrenesv.csv','vega_spav.csv','yeclav.csv']
nrowfile = 'bran_spv.csv'

with open(nrowfile, 'r') as csvfile:
    csv_reader = csv.reader(csvfile)
    row_count = sum(1 for row in csv_reader)
print(f"Number of rows in '{nrowfile}': {row_count}")

#csv_files = [os.path.join(script_directory, csv_file_path) for csv_file_path in spaincsv]
csv_files = spaincsv

spain_af_ref = [[] for _ in range(0, row_count)]  
spain_af_alt = [[] for _ in range(0, row_count)]  
spain_gt_het = [[] for _ in range(0, row_count)]  

for spaincsv in csv_files:
    zero_count = 0
    one_count = 0
    het_count = 0
    
    with open(spaincsv, 'r') as csvfile:
        csv_reader = csv.reader(csvfile, delimiter=',')
        
        for row in csv_reader:
            
            zero_count = 0
            one_count = 0
            het_count = 0
                        
            for entry in row[4:]:
                zero_count += entry.count("0/0") + entry.count("0|0")
                one_count += entry.count("1/1") + entry.count("1|1")
                het_count += entry.count("0|1") + entry.count("1|0") + entry.count("0/1") + entry.count("1/0")
            
            denominator = zero_count + one_count + het_count
            
            if denominator != 0:
                gthet = het_count / denominator
                af_ref = (2 * zero_count + het_count) / (2 * denominator)
                af_alt = (2 * one_count + het_count) / (2 * denominator)
            else:
                gthet = 0
                af_ref = 0
                af_alt = 0
                
            spain_af_ref[csv_reader.line_num - 1].append(af_ref)
            spain_af_alt[csv_reader.line_num - 1].append(af_alt)
            spain_gt_het[csv_reader.line_num - 1].append(gthet)

average_af_ref_per_row = [sum(row) / len(csv_files) for row in spain_af_ref]
average_af_alt_per_row = [sum(row) / len(csv_files) for row in spain_af_alt]
average_gt_het_per_row = [sum(row) / len(csv_files) for row in spain_gt_het]


scadcsv = ['nor1v.csv', 'nor2v.csv',  'nor4v.csv',  'nor5v.csv',  'nor6v.csv', 'nor7v.csv' , 'par_nor6v.csv', 's1v.csv', 's2v.csv', 'S3v.csv','S5v.csv', 'sve2v.csv', 'sve3v.csv', 'sve4v.csv']
#csv_files1 = [os.path.join(script_directory, csv_file_path1) for csv_file_path1 in scadcsv]
csv_files1 = scadcsv
scad_af_ref = [[] for _ in range(0, row_count)]  
scad_af_alt = [[] for _ in range(0, row_count)]  
scad_gt_het = [[] for _ in range(0, row_count)]  

for scadcsv in csv_files1:
    
    zero_count = 0
    one_count = 0
    het_count = 0
    
    with open(scadcsv, 'r') as csvfile1:
        csv_reader1 = csv.reader(csvfile1, delimiter=',')
        
        for row1 in csv_reader1:
            
            zero_count = 0
            one_count = 0
            het_count = 0
            
            for entry in row1[4:]:
                zero_count += entry.count("0/0") + entry.count("0|0")
                one_count += entry.count("1/1") + entry.count("1|1")
                het_count += entry.count("0|1") + entry.count("1|0") + entry.count("0/1") + entry.count("1/0")
            
            denominator = zero_count + one_count + het_count
            
            if denominator != 0:
                gthet = het_count / denominator
                af_ref = (2 * zero_count + het_count) / (2 * denominator)
                af_alt = (2 * one_count + het_count) / (2 * denominator)
            else:
                gthet = 0
                af_ref = 0
                af_alt = 0
                
            scad_af_ref[csv_reader1.line_num - 1].append(af_ref)
            scad_af_alt[csv_reader1.line_num - 1].append(af_alt)
            scad_gt_het[csv_reader1.line_num - 1].append(gthet)
            
average_gt_het_per_row1 = [sum(row1) / len(csv_files1) for row1 in scad_gt_het]
average_af_ref_per_row1 = [sum(row1) / len(csv_files1) for row1 in scad_af_ref]
average_af_alt_per_row1 = [sum(row1) / len(csv_files1) for row1 in scad_af_alt]

#for i, avg_af_ref in enumerate(average_af_ref_per_row1):
    #print(f"Average Af_ref for Row {i + 1}: {avg_af_ref}")

alpscsv = ['austriav.csv', 'bri_frv.csv', 'chv.csv',  'cr_frv.csv', 'D1_D2v.csv', 'D3v.csv', 'D4v.csv', 'Dotherv.csv', 'Fgal1v.csv', 'fr1_5v.csv', 'fr4v.csv','fr6v.csv','fr_Pv.csv','galiber_frv.csv', 'ganon_frv.csv', 'gv_frv.csv', 'polandv.csv', 'vallon_frv.csv','italymixedv.csv']
#csv_files2 = [os.path.join(script_directory, csv_file_path2) for csv_file_path2 in alpscsv]
csv_files2 = alpscsv

alps_af_ref = [[] for _ in range(0, row_count)]  
alps_af_alt = [[] for _ in range(0, row_count)]  
alps_gt_het = [[] for _ in range(0, row_count)] 

for alpscsv in csv_files2:
    
    zero_count = 0
    one_count = 0
    het_count = 0
    
    with open(alpscsv, 'r') as csvfile2:
        csv_reader2 = csv.reader(csvfile2, delimiter=',')
        
        for row2 in csv_reader2:
            
            zero_count = 0
            one_count = 0
            het_count = 0
            
            for entry in row2[4:]:
                zero_count += entry.count("0/0") + entry.count("0|0")
                one_count += entry.count("1/1") + entry.count("1|1")
                het_count += entry.count("0|1") + entry.count("1|0") + entry.count("0/1") + entry.count("1/0")
            
            denominator = zero_count + one_count + het_count
            
            if denominator != 0:
                gthet = het_count / denominator
                af_ref = (2 * zero_count + het_count) / (2 * denominator)
                af_alt = (2 * one_count + het_count) / (2 * denominator)
            else:
                gthet = 0
                af_ref = 0
                af_alt = 0
                
            alps_af_ref[csv_reader2.line_num - 1].append(af_ref)
            alps_af_alt[csv_reader2.line_num - 1].append(af_alt)
            alps_gt_het[csv_reader2.line_num - 1].append(gthet)

average_af_ref_per_row2 = [sum(row2) / len(csv_files2) for row2 in alps_af_ref]
average_af_alt_per_row2 = [sum(row2) / len(csv_files2) for row2 in alps_af_alt]
average_gt_het_per_row2 = [sum(row2) / len(csv_files2) for row2 in alps_gt_het]

wb = Workbook()
ws = wb.active

#ws.cell(row=1, column=1, value="Spain") #ws.cell(row=1, column=2, value="Scad") #ws.cell(row=1, column=3, value="Alps")

for i, avg_af_ref in enumerate(average_af_ref_per_row):
    ws.cell(row=i + 1, column=1, value=avg_af_ref)
for i, avg_af_ref in enumerate(average_af_ref_per_row1):
    ws.cell(row=i + 1, column=2, value=avg_af_ref)
for i, avg_af_ref in enumerate(average_af_ref_per_row2):
    ws.cell(row=i + 1, column=3, value=avg_af_ref)
 
wb.save("af_ref.xlsx")
wb.close()

wb1 = Workbook()
ws1 = wb1.active

for i, avg_af_alt in enumerate(average_af_alt_per_row):
    ws1.cell(row=i + 1, column=1, value=avg_af_alt)
for i, avg_af_alt in enumerate(average_af_alt_per_row1):
    ws1.cell(row=i + 1, column=2, value=avg_af_alt)
for i, avg_af_alt in enumerate(average_af_alt_per_row2):
    ws1.cell(row=i + 1, column=3, value=avg_af_alt)

wb1.save("af_alt.xlsx")
wb1.close()

wb2 = Workbook()
ws2 = wb2.active

for i, avg_gt_het in enumerate(average_gt_het_per_row):
    ws2.cell(row=i + 1, column=1, value=avg_gt_het)
for i, avg_gt_het in enumerate(average_gt_het_per_row1):
    ws2.cell(row=i + 1, column=2, value=avg_gt_het)
for i, avg_gt_het in enumerate(average_gt_het_per_row2):
    ws2.cell(row=i + 1, column=3, value=avg_gt_het)

wb2.save("gt_het.xlsx")
wb2.close()

